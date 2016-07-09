"""
Operations on how to store prices information into Excel file.

Because prices information get from JD is kind of stable, we cannot
store every price we get every time.

Therefore, we open a new sheet for each item and there would only be
2 columns in the sheet, the first column is TIME and the second is PRICES.
And, everytime, if the price we get is the same as the last 2 prices we got,
we just update the previous price's TIME information.

As a result, we could make this data collection space-friendly.

Besides, there should be a cell in the sheet to calculate the average price.
this average price should be average over time.

In order to implement this function, we could add a cell to count the times that
we try to get a new price.

All of above, below shows the prototype of the data collections.
        A       B       C                 D
1      $ITEM_NAME      AVERAGEPRICES     COUNT
2      TIME    PRICES   $AP               $C
3      $T1      $P1
4      $T2      $P2
5      $T3      $P3
...

$AP = ($AP*$C+$CP)/($C+1)
"""
import time
import logging

from openpyxl import Workbook, load_workbook


def init_worksheet(ws, item_name):
    ws['A1'] = item_name
    ws['A2'] = 'TIME'
    ws['B2'] = 'PRICES'
    ws['C1'] = 'AVERAGEPRICES'
    ws['C2'] = 0
    ws['D1'] = 'COUNT'
    ws['D2'] = 0
    

class ExcelOperation(object):
    """
    Toppest call should be:
        test = ExcelOperation(file_name, data_dict, time)
        test()
    The instance is a callable item
    """
    def __init__(self, file_name, data_dict, time):
        """
        :type file_name: string, the name of the excel file to be written.
        :type data_dict: dict, pairs of item_name and prices.
        :type time: Linux time stamp
        """
        self.file_name = file_name
        self.data_dict = data_dict
        self.time = time

    def check_and_init(self, wb):
        """
        Check whether the worksheet that needs to be operated exists.
        If not exists, create one and init it.
        """
        for item_name in self.data_dict.keys():
            name = item_name[:30]
            try:
                ws = wb[name]
            except KeyError:
                ws = wb.create_sheet(title=name)
                init_worksheet(ws, item_name)

    def create_and_init(self):
        """
        Create the EXCEL file with file_name name.
        Initialize the file with the right format.
        Cells to be Initialized: A1, A2, B2, C1, D1, D2

        :rtype: initialized workbook
        """
        logging.info("Create EXCEL file...")
        wb = Workbook()
        for item_name in self.data_dict.keys():
            name = item_name[:30]
            ws = wb.create_sheet(title=name)
            init_worksheet(ws, item_name)
        return wb

    def open(self):
        """
        If the file exists, open it and check it.
        If not exist, create a new file and initialize it.
        """
        try:
            wb = load_workbook(self.file_name)
            self.check_and_init(wb)
        except IOError:
            wb = self.create_and_init()
        return wb

    def insert_data(self, wb):
        """
        Insert new data into the EXCEL file and
        return the dict containing info about the
        item whose price has changed.
        """
        res = dict()
        for item_name,price in self.data_dict.iteritems():
            name = item_name[:30]
            ws = wb[name]
            average, count = ws['C2'], ws['D2']
            # insert one new row into the sheet
            ws = wb[name]
            # check whether the prices are the same
            if ws.max_row > 3:
                last_price = ws.cell(row=ws.max_row, column=2)
                last_last_price = ws.cell(row=ws.max_row-1, column=2)
                if last_price.value == price and last_last_price.value == price:
                    # these 3 time's price are the same
                    # only update the time
                    ws.cell(row=ws.max_row, column=1).value = self.time
                else:
                    # not the same, just append the data
                    data_append = {'A':self.time, 'B':price}
                    ws.append(data_append)
            else:
                # not enough data to compare, just append the data
                # after this append, the max_row of this wb has increased
                last_price = ws.cell(row=ws.max_row, column=2)
                data_append = {'A':self.time, 'B':price}
                ws.append(data_append)
            # update count and average
            average.value = (average.value * count.value + price) / (count.value + 1)
            count.value = count.value + 1
            # Price has changed
            if ws.max_row > 3:
                if price != last_price.value:
                    content = "from %.2f to %.2f" % (last_price.value, price)
                    res.update({item_name:content})
        return res


    def close(self, wb):
        """
        According the examples in the documentation of openpyxl,
        the last operation would always be save.
        """
        while True:
            try:
                wb.save(self.file_name)
                break
            except IOError:
                time.sleep(0.5)

    def __call__(self):
        """
        The toppest level of excel-operation.
        Insert data_dict into file_name.

        The diagram that how every method is called:
            open -+-EXIST-> check_and_init  -+-> insert_data -> close
                  +-NONE--> create_and_init -+
        """
        wb = self.open()
        changed = self.insert_data(wb)
        self.close(wb)
        return changed